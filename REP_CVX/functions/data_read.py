# -*- coding: utf-8 -*-


"""
data_read Module
==============================================================================
All the functions in this module are used to read the data and structure them
for the use in the main class.

Functions
----------
database:   Reading the database and making the DataFramse with MultiIndex

sens_info:  Extracting the sensitivities from the shock excel file and building
            the dircetory of sensitivity excels and storing the information
            in a dictionary to be used in the main class
"""
def database(path):
    '''
    

    Parameters
    ----------
    path : string
        specifies the directory of the database excel.

    Returns
    -------
    SUT : DataFrame
        Supply and Use Table with economic factor and satellite accounts.
    U : DataFrame
        Use flows.
    V : DataFrame
        Supply flows.
    Z : DataFrame
        Supply and Use flows.
    S : DataFrame
        Satellite account.
    Y : DataFrame
        Final demand.
    VA : DataFrame
        Economic factor flows.
    X : DataFrame
        Total production of Activities and Commodities.

    '''
    import pandas as pd
    
    SUT = pd.read_excel(path,header=[0,1,2,3],index_col=[0,1,2,3,4]).droplevel(level=4)
    
        # importing use (U), supply (V), supply-use together (Z) and satellite accounts (S)
    U = SUT.loc['Commodities','Activities']
    V = SUT.loc['Activities','Commodities']
    Z = SUT.loc[['Commodities','Activities'], ['Commodities','Activities']]
    S = SUT.loc['Satellite Accounts',['Commodities','Activities']]
        
         # importing the findal demand (Y) matrix   
    Y  = pd.DataFrame(SUT.loc[['Commodities','Activities'], 'Final Demand'].sum(axis=1),index=Z.index,columns=['Total final demand'])

        # computing total value added (VA) by importing factors of production (F),  margins as factor (F_M) if present in the database
    try: 

        VA   = SUT.loc[['Factors of production','Margins'], ['Commodities', 'Activities']].droplevel(level=4)

        
    except:
        VA = SUT.loc['Factors of production', ['Commodities','Activities']]

        
        # computing total production vector (X)
    X = pd.DataFrame(Y.sum(axis=1) + Z.sum(axis=1), index=Z.index, columns=['Total Production'])
        
    return SUT,U,V,Z,S,Y,VA,X

def sens_info(path):
    
    '''
    

    Parameters
    ----------
    path : string
        specifies the directory of the database excel.

    Returns
    -------
    directs : list
        A list of all the directories in which the excel files related to the
        shocks are stored.
        
    sensitivity_info : dictionary
        A dictionary which stores all the information related to the sensitivities


    '''
    
    import openpyxl
    import os
    import shutil
    import pandas as pd

    # Loading the excel file from the given path and loading the 'main' sheet
    myworkbook=openpyxl.load_workbook(path)
    worksheet= myworkbook.get_sheet_by_name('main')
    
    # Taking the number of rows to make loop on all the rows
    rows = worksheet.max_row
    
    # Identifies the columns. 
    # NOTE: The structre of the excel file should be exactly the same of the 
    # tutorial
    sens_col = 5
    val_col  = 3
    par_col  = 2
    min_col  = 6
    max_col  = 7
    stp_col  = 8
    mat_col  = 9
    
    # All the possible types of the shocks
    matrices = ['Y','S','VA','Z']
    
    # An empty dictionary to store the sensitivities information
    sensitivity_info = {}
    counter = 0
    
    # Making a loop around the rows to find the parameteres with sensitivity
    for row in range(rows):
        
        # Check if the sensitivity columns in the excel sheets specifies 'Yes' or not
        if worksheet.cell(row=row+1, column=sens_col).value == 'Yes':
            
            # if YES: it needs to store the following informatio:
                #1. matrices that are going to be affected
                #2. maximum, minimum and the step value of the sensitivity
                #3. the parmeter that is supposed to be changes
            
            # Free list to append the affected matrices by the sensitivity
            mat_list=[]
            
            # Inputs by user as a string representing the affected matrices
            mat_str = worksheet.cell(row=row+1, column=mat_col).value
            
            # Converting the string to a list
            mat_str=mat_str.split(',')
            for i in range(len(mat_str)):
            
                # As in the excel file the user needs to insert multiple matrices
                # (for example: A,VA) the inserted string should be splited and
                # converted to separated characters and if it is in acceptable
                # characters, it will be added to the list
                if mat_str[i] in matrices:
                    mat_list.append(mat_str[i])
                    
            
            
            # Storing the information in the dictionary
            sensitivity_info['{}'.format(counter)] = {'parameter':  worksheet.cell(row=row+1, column=par_col).value,
                                                      'minimum':    worksheet.cell(row=row+1, column=min_col).value,
                                                      'maximum':    worksheet.cell(row=row+1, column=max_col).value,
                                                      'step':       worksheet.cell(row=row+1, column=stp_col).value,
                                                      'row':        row+1,
                                                      'matrices':   mat_list}
            
            # Counting the number of sensitivities
            counter+=1                                                    
    
    print ('{} sensitivities are found'.format(counter))
    
    # If there is no sensitiviity, it raise an error
    if counter == 0: raise ValueError('No sensitivity found. Check if the main sheet has the right structure.')
    
    # A free list to append the directories
    directs=[]
    
    # For every set of sensitivity, a new folder will be created which contains
    # all the excel file for different steps of the sensitivity
    for i in range (counter):
        
        # check if the directory exists or not
        dir = os.path.join('sens_{}'.format(sensitivity_info[str(i)]['parameter']))
        
        # if doesnt exist, make it.
        if not os.path.exists(dir):
            os.mkdir(dir)
            
        # if exists, delete it and make a new one
        else:
            shutil.rmtree(dir)
            os.mkdir(dir)
        
        # store the directory into the list
        directs.append(dir)
        
        # reading the information of every set of sensitivity
        s_min = sensitivity_info[str(i)]['minimum']
        s_max = sensitivity_info[str(i)]['maximum']
        step  = sensitivity_info[str(i)]['step']
        row   = sensitivity_info[str(i)]['row']
        
        # a loop for implementing the changes in the excel file for the range
        # of (s_min,s_max,step)
        while s_min <= s_max:
            
            # loading the main excel file
            myworkbook=openpyxl.load_workbook(path)

            # loading the main sheet in which all the calculations are done!
            worksheet= myworkbook.get_sheet_by_name('main')
            
            # implement the change
            worksheet.cell(row=row, column=val_col).value = s_min
            
            # save the new excel file
            name = directs[i] + '\case_{}.xlsx'.format(s_min)
            myworkbook.save(name)
            
            s_min+=step   


            
    return directs,sensitivity_info


def sensitivity_take(variable,scenario,results,aggregation,level,indicator,m_unit,unit,title,rational,indeces):
    
    from REP_CVX.functions.check import unit_check
    from REP_CVX.functions.check import unit_converter
    from REP_CVX.functions.check import level_check
    from REP_CVX.functions.check import indic_check

       
    
    # default unit
    if unit == 'default':
        unit = m_unit
        
    tit,level = level_check(level)
    conversion = unit_converter(m_unit,unit_check(unit))
    if  indicator != None: indicator   = indic_check (indicator,list(indeces['S_ind'].get_level_values(1)))
    
    # this dictionary contains the full name of the variables
    var_name = {'X': 'Production' , 'VA': 'Value Added', 'S': '{} change'.format(indicator) }    
    
    try:    data = results['sensitivity_{}'.format(scenario)]
    except: raise ValueError('sensitivity_{} does not exist in results.'.format(scenario))

    if aggregation: var = '{}_agg'.format(variable)

    
    sens_all=[]
    for key, value in results['sensitivity_{}'.format(scenario)].items(): 
        if key != 'information':
            if variable == 'X':
                sens_all.append(data[key][var].loc[level].values*conversion-results['baseline'][var].loc[level].values*conversion)
            if variable == 'VA':
                sens_all.append(data[key][var][level].sum(axis=rational)*conversion-results['baseline'][var][level].sum(axis=rational)*conversion)
            if variable == 'S':
                sens_all.append(data[key][var][level].sum(axis=rational)*conversion-results['baseline'][var][level].sum(axis=rational)*conversion)
                

                
         
    for i in range(len(sens_all)):
        sens_all[i]=sens_all[i].ravel()

    # Reshaping the data into appropriate form
    sen_to_plt=[]
    for i in range(len(sens_all[0])):
        dt=[]
        for j in range(len(sens_all)):
            dt.append(sens_all[j][i])
        sen_to_plt.append(dt)

    
    if variable == 'X':    index = data[key][var].loc[level].index
    elif variable == 'VA':
        if rational == 0:
            index = data[key][var][level].columns
        elif rational == 1:
            index = data[key][var][level].index
    elif variable == 'S':
        index = data[key][var][level].columns 
        unit = ''
    
    if len(level) != 2:
        try: index = index.get_level_values(1)
        except: index = index

    
    if title == 'default':
        title = '{}{}'.format(var_name[variable],tit)
    
    legend = '{}, \n range= {} to {}'.format(results['sensitivity_{}'.format(scenario)]['information']['parameter'],results['sensitivity_{}'.format(scenario)]['information']['minimum'],results['sensitivity_{}'.format(scenario)]['information']['maximum'])
    return sen_to_plt,index,title,legend,unit













        